"""
recognition_program/render.py

Generates a tiered volunteer recognition program workbook as an Excel (.xlsx) file.

Sheet 1 — Program Design: tiered recognition table with milestone types,
           recognition examples, and estimated costs.
Sheet 2 — Tracking Template: empty volunteer roster for tracking progress
           toward milestones.

All libraries are pre-installed in Anthropic's code-execution sandbox.
"""

import os
import time
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

NAVY = "1A476B"
STEEL = "2E729E"
LIGHT_BLUE = "D6E8F5"
LIGHT_GOLD = "FFF8DC"
LIGHT_GRAY = "F5F5F5"
WHITE = "FFFFFF"
DARK_GRAY = "333333"
MED_GRAY = "666666"

def _header_font(bold=True, color=WHITE, size=11):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def _body_font(bold=False, color=DARK_GRAY, size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def _header_fill(hex_color=NAVY):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def _row_fill(hex_color=WHITE):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def _thin_border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------
# Recognition tier builder
# ---------------------------------------------------------------------------

def _build_tiers(milestone_types: list, program_focus: str, recognition_budget: Optional[str]) -> list:
    """
    Build the recognition tier rows from milestone_types input.

    milestone_types is a list of dicts, e.g.:
      [
        {"type": "hours", "values": [25, 50, 100]},
        {"type": "tenure", "values": ["6mo", "1yr", "2yr"]},
        {"type": "special_contribution"}
      ]

    Returns list of dicts: {tier, milestone, recognition_type, examples, cost}
    """
    rows = []
    tier_counter = 1
    budget_constrained = recognition_budget and recognition_budget.lower() not in ("none", "0", "$0", "n/a")

    for milestone_obj in milestone_types:
        mtype = milestone_obj.get("type", "")
        values = milestone_obj.get("values", [])

        if mtype == "hours":
            for val in values:
                tier_name = f"Tier {tier_counter}"
                milestone = f"{val} volunteer hours"

                if val <= 25:
                    rec_type = "Digital Recognition"
                    examples = "Personalized thank-you email from ED; name on volunteer Wall of Honor (digital or physical); social media shoutout (with permission)"
                    cost = "$0 – $5"
                elif val <= 75:
                    rec_type = "Public Recognition + Tangible Gift"
                    examples = "Certificate of appreciation; branded volunteer t-shirt or tote; acknowledgment at all-staff or board meeting"
                    cost = "$15 – $30" if budget_constrained else "Varies by budget"
                else:
                    rec_type = "Milestone Award + Elevated Access"
                    examples = "Framed certificate; branded gear (jacket, water bottle); invitation to leadership roundtable or site tour; personal note from ED"
                    cost = "$40 – $75" if budget_constrained else "Varies by budget"

                rows.append({
                    "tier": tier_name,
                    "milestone": milestone,
                    "recognition_type": rec_type,
                    "examples": examples,
                    "cost": cost,
                })
                tier_counter += 1

        elif mtype == "tenure":
            for val in values:
                tier_name = f"Tier {tier_counter}"
                milestone = f"{val} of service"

                if "6" in str(val):
                    rec_type = "Early Recognition"
                    examples = "Handwritten thank-you card from supervisor; feature in newsletter or email update; small branded gift (pin, sticker pack)"
                    cost = "$5 – $15"
                elif "1yr" in str(val) or "1 yr" in str(val) or val in ("1yr", "12mo", "1 year"):
                    rec_type = "Anniversary Recognition"
                    examples = "Engraved or personalized keepsake; celebration at volunteer appreciation event; peer nomination letter in personnel file"
                    cost = "$20 – $40"
                else:
                    rec_type = "Long-Service Award"
                    examples = "Dedicated award (trophy, plaque, or framed artwork); named recognition in annual report; special role or advisory designation"
                    cost = "$50 – $100"

                rows.append({
                    "tier": tier_name,
                    "milestone": milestone,
                    "recognition_type": rec_type,
                    "examples": examples,
                    "cost": cost,
                })
                tier_counter += 1

        elif mtype == "special_contribution":
            tier_name = f"Tier {tier_counter} (Special)"
            milestone = "Exceptional / above-and-beyond contribution"
            rec_type = "Spotlight Award (Discretionary)"
            examples = (
                f"Nomination-based award presented at {program_focus or 'volunteer'} events; "
                "choice of recognition gift from approved list; "
                "profile story in donor newsletter or social media; "
                "personal call or visit from executive leadership"
            )
            cost = "$25 – $75 (discretionary)"
            rows.append({
                "tier": tier_name,
                "milestone": milestone,
                "recognition_type": rec_type,
                "examples": examples,
                "cost": cost,
            })
            tier_counter += 1

        else:
            # Generic fallback for unknown types
            tier_name = f"Tier {tier_counter}"
            milestone = str(mtype)
            rec_type = "Custom Recognition"
            examples = "To be defined by HR coordinator"
            cost = "TBD"
            rows.append({
                "tier": tier_name,
                "milestone": milestone,
                "recognition_type": rec_type,
                "examples": examples,
                "cost": cost,
            })
            tier_counter += 1

    return rows


# ---------------------------------------------------------------------------
# Core render function
# ---------------------------------------------------------------------------

def render(
    org_name: str,
    milestone_types: list,
    recognition_budget: Optional[str] = None,
    program_focus: Optional[str] = None,
    output_dir: str = "/mnt/user-data/outputs",
) -> str:
    """
    Generate a volunteer recognition program Excel workbook.

    Parameters
    ----------
    org_name : str
        Name of the organization
    milestone_types : list
        List of milestone objects, each with "type" and optional "values".
        Supported types: "hours", "tenure", "special_contribution"
        Example: [
            {"type": "hours", "values": [25, 50, 100]},
            {"type": "tenure", "values": ["6mo", "1yr", "2yr"]},
            {"type": "special_contribution"}
        ]
    recognition_budget : str, optional
        Dollar amount or "none". Used to calibrate cost estimates.
    program_focus : str, optional
        Program or community focus, e.g. "youth services". Used in award language.
    output_dir : str
        Directory to save the output file

    Returns
    -------
    str
        Absolute path to the generated .xlsx file
    """
    # --- Validate required inputs ---
    if not org_name:
        raise ValueError("org_name is required")
    if not milestone_types:
        raise ValueError("milestone_types is required and must be a non-empty list")

    program_focus = program_focus or "community"

    wb = Workbook()

    # =========================================================
    # SHEET 1: Program Design
    # =========================================================
    ws1 = wb.active
    ws1.title = "Program Design"

    # Title block
    ws1.merge_cells("A1:E1")
    title_cell = ws1["A1"]
    title_cell.value = f"{org_name} — Volunteer Recognition Program"
    title_cell.font = Font(bold=True, size=14, color=NAVY, name="Calibri")
    title_cell.fill = _row_fill(LIGHT_BLUE)
    title_cell.alignment = _center()
    ws1.row_dimensions[1].height = 28

    ws1.merge_cells("A2:E2")
    sub_cell = ws1["A2"]
    sub_cell.value = (
        f"Program Focus: {program_focus.title()}  |  "
        f"Budget: {recognition_budget or 'TBD'}  |  "
        f"Generated: {time.strftime('%Y-%m-%d')}"
    )
    sub_cell.font = Font(italic=True, size=10, color=MED_GRAY, name="Calibri")
    sub_cell.fill = _row_fill(LIGHT_BLUE)
    sub_cell.alignment = _center()
    ws1.row_dimensions[2].height = 18

    # Blank spacer row
    ws1.row_dimensions[3].height = 6

    # Column headers — row 4
    headers = ["Tier", "Milestone", "Recognition Type", "Examples", "Estimated Cost"]
    col_widths = [12, 26, 22, 52, 18]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws1.cell(row=4, column=col_idx, value=header)
        cell.font = _header_font()
        cell.fill = _header_fill(NAVY)
        cell.alignment = _center()
        cell.border = _thin_border()
        ws1.column_dimensions[get_column_letter(col_idx)].width = width

    ws1.row_dimensions[4].height = 22

    # Freeze header row
    ws1.freeze_panes = "A5"

    # Data rows
    tiers = _build_tiers(milestone_types, program_focus, recognition_budget)

    for row_num, tier in enumerate(tiers, start=5):
        bg = LIGHT_GOLD if "Special" in tier["tier"] else (LIGHT_BLUE if row_num % 2 == 0 else WHITE)

        values = [
            tier["tier"],
            tier["milestone"],
            tier["recognition_type"],
            tier["examples"],
            tier["cost"],
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws1.cell(row=row_num, column=col_idx, value=value)
            cell.font = _body_font(bold=(col_idx == 1))
            cell.fill = _row_fill(bg)
            cell.border = _thin_border()
            cell.alignment = _left() if col_idx in (2, 3, 4) else _center()

        ws1.row_dimensions[row_num].height = 52

    # Implementation notes below the table
    notes_row = len(tiers) + 6
    ws1.merge_cells(f"A{notes_row}:E{notes_row}")
    notes_header = ws1.cell(row=notes_row, column=1, value="Implementation Notes")
    notes_header.font = Font(bold=True, size=10, color=NAVY, name="Calibri")

    note_lines = [
        "1. Designate one owner per milestone type (e.g., Volunteer Coordinator owns hours milestones; Program Manager owns special contributions).",
        "2. Recognize milestones within 30 days of achievement — delayed recognition loses impact.",
        "3. Always ask for the volunteer's communication preference before social media shoutouts.",
        "4. Review cost estimates annually; adjust to actual budget allocation.",
        "5. For budget-constrained programs: prioritize handwritten notes and public verbal recognition — both have high impact at zero cost.",
    ]

    for i, note in enumerate(note_lines, start=notes_row + 1):
        ws1.merge_cells(f"A{i}:E{i}")
        cell = ws1.cell(row=i, column=1, value=note)
        cell.font = _body_font(size=9, color=MED_GRAY)
        cell.alignment = _left()
        ws1.row_dimensions[i].height = 16

    # =========================================================
    # SHEET 2: Tracking Template
    # =========================================================
    ws2 = wb.create_sheet(title="Tracking Template")

    # Title
    ws2.merge_cells("A1:G1")
    t2 = ws2["A1"]
    t2.value = f"{org_name} — Volunteer Recognition Tracking Roster"
    t2.font = Font(bold=True, size=13, color=NAVY, name="Calibri")
    t2.fill = _row_fill(LIGHT_BLUE)
    t2.alignment = _center()
    ws2.row_dimensions[1].height = 26

    ws2.merge_cells("A2:G2")
    s2 = ws2["A2"]
    s2.value = "Instructions: Add one row per active volunteer. Update after each milestone is reached. 'Next Milestone' and 'Owner' columns drive accountability."
    s2.font = Font(italic=True, size=9, color=MED_GRAY, name="Calibri")
    s2.fill = _row_fill(LIGHT_BLUE)
    s2.alignment = _left()
    ws2.row_dimensions[2].height = 18

    # Headers — row 4
    track_headers = [
        "Volunteer Name",
        "Start Date",
        "Hours Logged",
        "Last Recognition Date",
        "Recognition Given",
        "Next Milestone",
        "Owner",
    ]
    track_widths = [24, 14, 16, 22, 28, 22, 18]

    for col_idx, (header, width) in enumerate(zip(track_headers, track_widths), start=1):
        cell = ws2.cell(row=4, column=col_idx, value=header)
        cell.font = _header_font()
        cell.fill = _header_fill(STEEL)
        cell.alignment = _center()
        cell.border = _thin_border()
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    ws2.row_dimensions[4].height = 22
    ws2.freeze_panes = "A5"

    # 20 blank data rows with light banding
    for row_num in range(5, 25):
        bg = LIGHT_GRAY if row_num % 2 == 0 else WHITE
        for col_idx in range(1, len(track_headers) + 1):
            cell = ws2.cell(row=row_num, column=col_idx, value="")
            cell.fill = _row_fill(bg)
            cell.border = _thin_border()
            cell.alignment = _left()
            cell.font = _body_font()
        ws2.row_dimensions[row_num].height = 18

    # --- Save ---
    os.makedirs(output_dir, exist_ok=True)
    timestamp = int(time.time())
    filename = f"recognition_program_{timestamp}.xlsx"
    out_path = os.path.join(output_dir, filename)
    wb.save(out_path)

    return out_path
