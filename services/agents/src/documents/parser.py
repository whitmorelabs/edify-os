"""Document text extraction for various file formats.

Dependencies (add to requirements.txt if not already present):
  - PyMuPDF          (import fitz)      -- PDF extraction
  - python-docx      (import docx)      -- DOCX extraction
  - openpyxl                            -- XLSX/XLS extraction
  - chardet                             -- charset detection (optional, graceful fallback)
"""

from __future__ import annotations

import csv
import io
import logging
import re

logger = logging.getLogger(__name__)

MAX_TEXT_LENGTH = 100_000  # characters -- truncate beyond this


class DocumentParser:
    """Extract plain text from PDF, DOCX, TXT, MD, CSV, XLS, and XLSX files.

    Usage::

        parser = DocumentParser()
        text = parser.parse("/tmp/report.pdf", "pdf")
    """

    def parse(self, file_path: str, file_type: str) -> str:
        """Extract text from *file_path*.

        Parameters
        ----------
        file_path:
            Absolute path to the file on disk.
        file_type:
            Lower-cased extension without dot, e.g. ``"pdf"``, ``"docx"``.

        Returns
        -------
        str
            Cleaned, normalised text.  Truncated to MAX_TEXT_LENGTH characters
            with a warning when the document is very large.
        """
        file_type = file_type.lower().lstrip(".")

        dispatch = {
            "pdf": self._parse_pdf,
            "docx": self._parse_docx,
            "doc": self._parse_docx,  # fallback -- python-docx handles most .doc too
            "txt": self._parse_text,
            "md": self._parse_text,
            "csv": self._parse_csv,
            "xlsx": self._parse_spreadsheet,
            "xls": self._parse_spreadsheet,
        }

        handler = dispatch.get(file_type)
        if handler is None:
            raise ValueError(f"Unsupported file type: {file_type!r}")

        try:
            raw = handler(file_path)
        except Exception as exc:
            raise RuntimeError(
                f"Failed to parse {file_type!r} file '{file_path}': {exc}"
            ) from exc

        cleaned = self._clean(raw)

        if len(cleaned) > MAX_TEXT_LENGTH:
            logger.warning(
                "Document '%s' truncated from %d to %d characters.",
                file_path,
                len(cleaned),
                MAX_TEXT_LENGTH,
            )
            cleaned = cleaned[:MAX_TEXT_LENGTH]

        return cleaned

    # ------------------------------------------------------------------
    # Format-specific handlers
    # ------------------------------------------------------------------

    def _parse_pdf(self, file_path: str) -> str:
        """Extract text from a PDF using PyMuPDF (fitz)."""
        try:
            import fitz  # PyMuPDF
        except ImportError as exc:
            raise ImportError(
                "PyMuPDF is required for PDF parsing. Install it with: pip install PyMuPDF"
            ) from exc

        parts: list[str] = []
        with fitz.open(file_path) as doc:
            for page_num, page in enumerate(doc, start=1):
                text = page.get_text()
                if text.strip():
                    parts.append(f"[Page {page_num}]\n{text}")

        return "\n\n".join(parts)

    def _parse_docx(self, file_path: str) -> str:
        """Extract text from a DOCX file using python-docx."""
        try:
            import docx  # python-docx
        except ImportError as exc:
            raise ImportError(
                "python-docx is required for DOCX parsing. Install it with: pip install python-docx"
            ) from exc

        document = docx.Document(file_path)
        paragraphs = [p.text for p in document.paragraphs if p.text.strip()]
        return "\n\n".join(paragraphs)

    def _parse_text(self, file_path: str) -> str:
        """Read plain text or Markdown with charset detection."""
        # Try UTF-8 first; fall back to latin-1 which never fails.
        for encoding in ("utf-8", "utf-8-sig", "latin-1"):
            try:
                with open(file_path, "r", encoding=encoding) as fh:
                    return fh.read()
            except UnicodeDecodeError:
                continue

        # Should never reach here because latin-1 is lossless.
        raise RuntimeError(f"Could not decode text file: {file_path}")

    def _parse_csv(self, file_path: str) -> str:
        """Read a CSV and render each row as a readable line."""
        rows: list[str] = []

        for encoding in ("utf-8", "utf-8-sig", "latin-1"):
            try:
                with open(file_path, "r", encoding=encoding, newline="") as fh:
                    reader = csv.reader(fh)
                    headers: list[str] | None = None
                    for i, row in enumerate(reader):
                        if i == 0:
                            headers = row
                            rows.append("Headers: " + " | ".join(row))
                        else:
                            if headers:
                                # key: value pairs for readability
                                pairs = [
                                    f"{h}: {v}"
                                    for h, v in zip(headers, row)
                                    if v.strip()
                                ]
                                rows.append("; ".join(pairs))
                            else:
                                rows.append(" | ".join(row))
                break  # success
            except UnicodeDecodeError:
                rows = []
                continue

        return "\n".join(rows)

    def _parse_spreadsheet(self, file_path: str) -> str:
        """Extract cell values from XLSX (and basic XLS) using openpyxl."""
        try:
            import openpyxl
        except ImportError as exc:
            raise ImportError(
                "openpyxl is required for spreadsheet parsing. Install it with: pip install openpyxl"
            ) from exc

        # openpyxl handles .xlsx natively.  For legacy .xls files, it will
        # raise InvalidFileException -- callers should convert first or accept
        # the degraded error message.
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        parts: list[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_rows: list[str] = [f"[Sheet: {sheet_name}]"]

            headers: list[str] | None = None
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                str_cells = [str(c) if c is not None else "" for c in row]
                if all(v == "" for v in str_cells):
                    continue  # skip empty rows

                if i == 0:
                    headers = str_cells
                    sheet_rows.append("Headers: " + " | ".join(str_cells))
                else:
                    if headers:
                        pairs = [
                            f"{h}: {v}"
                            for h, v in zip(headers, str_cells)
                            if v.strip()
                        ]
                        sheet_rows.append("; ".join(pairs))
                    else:
                        sheet_rows.append(" | ".join(str_cells))

            parts.append("\n".join(sheet_rows))

        wb.close()
        return "\n\n".join(parts)

    # ------------------------------------------------------------------
    # Text cleaning
    # ------------------------------------------------------------------

    def _clean(self, text: str) -> str:
        """Normalise whitespace and line breaks."""
        # Collapse runs of 3+ blank lines into two
        text = re.sub(r"\n{3,}", "\n\n", text)
        # Strip trailing whitespace from each line
        lines = [line.rstrip() for line in text.splitlines()]
        return "\n".join(lines).strip()
